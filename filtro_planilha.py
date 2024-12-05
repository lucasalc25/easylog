import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Carregar a planilha
file_path = './planilhas/rptPresencaFaltaPeriodo.xls'  # Substitua pelo nome correto do arquivo
sheet_name = 'Sheet'  # Substitua pelo nome correto da aba, caso necessário
df = pd.read_excel(file_path, sheet_name=sheet_name, header=3)

# Remover linhas onde a coluna 'Curso' tenha "Annual Book - Multimídia"
df = df[df['Curso'] != 'Annual Book - Multimídia']

# Manter apenas as colunas "aluno", "tel residencial" e "celular"
df = df[['Aluno', 'Educador', 'Falta', 'Celular']]

# Remover linhas onde a coluna "Falta" tem valor 0
df = df[df['Falta'] != 0]

# Remover duplicatas na coluna "Celular"
df = df.drop_duplicates(subset='Celular', keep='first')

# Remover colunas "B", "E" e "F"
cols_to_remove = ['B', 'E', 'F']
df = df.drop(columns=cols_to_remove, errors='ignore')

# Preencher células vazias na coluna "Celular" (C) com valores da coluna "Tel Residencial" (B)
df['Celular'] = df['Celular'].fillna(df['Tel Residencial'])

# Remove a coluna "B"
df = df.drop('Tel Residencial', axis=1, errors='ignore')

# Remover linhas duplicadas
df = df.drop_duplicates()

# Salvar as alterações para ajustar a largura das colunas
temp_file_path = './planilhas/faltas_temporario.xlsx'
df.to_excel(temp_file_path, index=False)

# Ajustar larguras das colunas com openpyxl
wb = load_workbook(temp_file_path)
ws = wb.active

# Ajustar a largura das colunas especificadas
column_widths = {
    'A': 35,  # Coluna A
    'B': 15,   # Coluna B
    'C': 15   # Coluna C
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Salvar o arquivo final
output_file_path = './planilhas/faltas_filtradas.xlsx'
wb.save(output_file_path)

print(f"Arquivo salvo como {output_file_path}")